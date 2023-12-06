import streamlit as st
import pandas as pd
import openpyxl
from matplotlib import pyplot as plt
import plotly.express as px
from PIL import Image

def app():
    excel_file = r"C:\Users\tejas\PycharmProjects\BtechPr\PrApp\Mastersheet_1.xlsx"

    df = pd.read_excel(excel_file)

    df.dropna(inplace=True)
    print(df)
    st.sidebar.title("Please Filter Here:")
    st.sidebar.title(" ")
    project = st.sidebar.multiselect(
        "Select the Project:",
        options=df["Project_Id"].unique(),
        default=df["Project_Id"].unique()
    )

    st.sidebar.title(" ")
    Task = st.sidebar.multiselect(
        "Select the Task Status:",
        options=df["Task_Status"].unique(),
        default=df["Task_Status"].unique()
    )

    df_selection = df.query(
        "Project_Id == @project & Task_Status ==@Task"
    )
    print(project)

    # TOP KPI's
    st.markdown("""---""")
    total_projects = int(len(project))
    Status_of_task = int(df["Task_Status"].count())

    Employee = int(df["Employee_Id"].count())

    left_column, right_column = st.columns(2)
    with left_column:
        st.subheader(":star:Total Projects you working on:")


    with right_column:
        st.subheader(":star:Total Employees Working Under You:")

    l1,l2,m1,m2=st.columns(4)
    with l2:
        st.title(f"{total_projects}")
    with m2:
        st.title(Employee)
    st.markdown("""---""")

    left,mid,right=st.columns(3)
    with mid:
        st.dataframe(df_selection[["Project_Id","Employee_Id","Task_Name","Task_Status","N0_OF_REMINDERS"]])

    df2=df_selection[["Project_Id","Employee_Id","Task_Name","Task_Status"]]
    print(df2)


    #to see data of individual employee
    wb = openpyxl.load_workbook('Mastersheet_1.xlsx')
    sheet = wb['Sheet1']
    employees = {}
    manager = {}
    managers_and_employees = {}

    for r in range(2, sheet.max_row + 1):
        designation = sheet.cell(row=r, column=2).value
        taskStatus = sheet.cell(row=r, column=7).value
        if designation == 'Employee':
            email = sheet.cell(row=r, column=9).value
            projectId = sheet.cell(row=r, column=1).value
            taskId = sheet.cell(row=r, column=4).value
            empId = sheet.cell(row=r, column=3).value
            taskStatus = sheet.cell(row=r, column=7).value
            dueDate = sheet.cell(row=r, column=6).value
            reminders = sheet.cell(row=r, column=8).value

            if email not in employees:
                employees[email] = []

            employees[email].append({
                'empId': empId,
                'projectId': projectId,
                'taskId': taskId,
                'dueDate': dueDate,
                'taskStatus': taskStatus,
                'reminders': reminders

            })

        elif designation == 'Manager':
            email = sheet.cell(row=r, column=9).value
            projectId = sheet.cell(row=r, column=1).value

            if email not in manager:
                manager[email] = []
            manager[email].append(projectId)

    # Populate dictionary based on existing data
    for employee_mail, employee_info_list in employees.items():
        for emp_info in employee_info_list:
            projectId = emp_info['projectId']

            for email, proj_Id in manager.items():
                if projectId in proj_Id:
                    manager_email = email

                    if manager_email not in managers_and_employees:
                        managers_and_employees[manager_email] = []

                    managers_and_employees[manager_email].append({
                        'empId': emp_info['empId'],
                        'projectId': emp_info['projectId'],
                        'taskId': emp_info['taskId'],
                        'taskStatus': emp_info['taskStatus'],
                        'dueDate': emp_info['dueDate'],
                        'reminders': emp_info['reminders'],
                    })
                    break


    st.markdown("""---""")
    st.header("See Employees Individual  Details:")
    l1, l2, l3 = st.columns([1, 1, 1])
    for employee_mail, employee_info_list in employees.items():
        cnt = 0
        i = 1

        for emp_info in employee_info_list:
            empId = emp_info['empId']
            dueDate = emp_info['dueDate']
            taskId = emp_info['taskId']
            reminders = emp_info['reminders']
            taskStatus = emp_info['taskStatus']
            projectId = emp_info['projectId']
            if i == 1:
                with l1:
                    exp = st.expander(label=f"Emp Id:{empId}")
                    exp.write(f"**Project Id:** {projectId}")
                    exp.write(f"**TaskID:** {taskId} ")
                    exp.write(f"**Due Date:** {dueDate} ")
                    exp.write(f"**Reminders:** {reminders}, ")
                    exp.subheader(f"{taskStatus}")
                cnt = cnt + 1
                if cnt == 1:
                    i = 2
                    cnt = 0

            elif i == 2:
                with l2:
                    exp = st.expander(label=f"Emp Id:{empId}")
                    exp.write(f"Project Id:{projectId}")
                    exp.write(f"TaskID: {taskId} ")
                    exp.write(f"Due Date: {dueDate} ")
                    exp.write(f"Reminders: {reminders}, ")
                    exp.subheader(f"{taskStatus}")

                cnt = cnt + 1
                if cnt == 1:
                    i = 3
                    cnt = 0

            elif i == 3:
                with l3:
                    exp = st.expander(label=f"Emp Id:{empId}")
                    exp.write(f"Project Id:{projectId}")
                    exp.write(f"TaskID: {taskId} ")
                    exp.write(f"Due Date: {dueDate} ")
                    exp.write(f"Reminders: {reminders}, ")
                    exp.subheader(f"{taskStatus}")
                cnt = cnt + 1
                if cnt == 1:
                    cnt = 0
    st.markdown("""---""")





